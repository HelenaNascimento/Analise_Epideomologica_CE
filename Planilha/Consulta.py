import pandas as pd
import numpy as np
import pyodbc
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURAÇÕES
# ============================================================

CAMINHO_PLANILHA = r"C:\Workstation\ScriptsSQL\Planilha\precos_planilha.xlsx"
CAMINHO_SAIDA = r"C:\Workstation\ScriptsSQL\Planilha\divergencias_precos.xlsx"

NOME_ABA_PLANILHA = 0  # Pode ser 0 para primeira aba ou o nome da aba. Ex: "Planilha1"

# Colunas da planilha
COLUNA_SKU_PLANILHA = "SKU"
COLUNA_PRODUTO_PLANILHA = "Produto"
COLUNA_PRECO_PLANILHA = "Preco_Planilha"

# Tolerância para diferença de preço
# Exemplo: 0.01 permite diferença de até 1 centavo
TOLERANCIA = 0.01


# ============================================================
# CONFIGURAÇÃO SQL SERVER
# ============================================================

SERVIDOR = "192.168.100.6"
BANCO = "DMD"
USUARIO = "sainfarma"
SENHA = "SAInfarma2022@"

QUERY_BANCO = """
SELECT 
	distinct
	PR.Cod_EAN as SKU,
	replace(convert(decimal(10,2), es.Prc_Venda), '.', ',') as Preco_Banco
	FROM PRODU PR
	JOIN PRXES ES ON PR.CODIGO = ES.COD_PRODUT
	JOIN PCXPR PC ON ES.Cod_Produt = PC.Cod_Produt
WHERE Cod_Estabe = 1
AND	COD_FABRICANTE = 321
AND PR.Flag_ImprClassif1 <> 'N'
and tipo = '00'
and pc.Id_PolCom = 2873
"""


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def normalizar_sku(valor):
    """
    Normaliza código/SKU vindo do Excel ou banco.
    Evita problema do Excel transformar 12345 em 12345.0.
    """
    if pd.isna(valor):
        return np.nan

    valor = str(valor).strip()

    if valor == "":
        return np.nan

    if valor.endswith(".0"):
        valor = valor[:-2]

    return valor


def normalizar_preco(valor):
    """
    Converte preços em formato brasileiro para número.

    Exemplos aceitos:
    10,50
    1.234,56
    R$ 10,50
    10.50
    100
    """
    if pd.isna(valor):
        return np.nan

    valor = str(valor).strip()

    if valor == "":
        return np.nan

    valor = valor.replace("R$", "")
    valor = valor.replace("r$", "")
    valor = valor.replace(" ", "")

    # Formato brasileiro: 1.234,56
    if "," in valor:
        valor = valor.replace(".", "")
        valor = valor.replace(",", ".")

    try:
        return float(
            Decimal(valor).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP
            )
        )
    except (InvalidOperation, ValueError):
        return np.nan


def validar_colunas(df, colunas_obrigatorias, origem):
    colunas_ausentes = [
        coluna for coluna in colunas_obrigatorias
        if coluna not in df.columns
    ]

    if colunas_ausentes:
        raise Exception(
            f"As seguintes colunas não foram encontradas em {origem}: "
            f"{', '.join(colunas_ausentes)}"
        )


def classificar_linha(row):
    if row["_merge"] == "left_only":
        return "Produto existe na planilha, mas não existe no banco"

    if row["_merge"] == "right_only":
        return "Produto existe no banco, mas não existe na planilha"

    preco_planilha = row["Preco_Planilha"]
    preco_banco = row["Preco_Banco"]

    if pd.isna(preco_planilha):
        return "Preço vazio ou inválido na planilha"

    if pd.isna(preco_banco):
        return "Preço vazio ou inválido no banco"

    diferenca = abs(preco_planilha - preco_banco)

    if diferenca > TOLERANCIA:
        return "Preço divergente"

    return "Preço igual"


def formatar_excel(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)

    fill_header = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    font_header = Font(
        color="FFFFFF",
        bold=True
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                valor = cell.value
                if valor is not None:
                    max_length = max(max_length, len(str(valor)))

            ws.column_dimensions[col_letter].width = min(max_length + 2, 45)

        # Formatar colunas numéricas
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter in ["C", "D", "E", "F", "G", "H", "I", "J"]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'

    wb.save(caminho_arquivo)


# ============================================================
# INÍCIO DO PROCESSAMENTO
# ============================================================

print("Iniciando comparação de preços...")

arquivo_planilha = Path(CAMINHO_PLANILHA)

if not arquivo_planilha.exists():
    raise FileNotFoundError(f"Planilha não encontrada: {CAMINHO_PLANILHA}")


# ============================================================
# LER PLANILHA
# ============================================================

print("Lendo planilha...")

df_planilha = pd.read_excel(
    CAMINHO_PLANILHA,
    sheet_name=NOME_ABA_PLANILHA,
    dtype=str
)

validar_colunas(
    df_planilha,
    [COLUNA_SKU_PLANILHA, COLUNA_PRECO_PLANILHA],
    "planilha"
)

# Renomeia colunas da planilha para o padrão do script
renomear_colunas = {
    COLUNA_SKU_PLANILHA: "SKU",
    COLUNA_PRECO_PLANILHA: "Preco_Planilha"
}

if COLUNA_PRODUTO_PLANILHA in df_planilha.columns:
    renomear_colunas[COLUNA_PRODUTO_PLANILHA] = "Produto_Planilha"

df_planilha = df_planilha.rename(columns=renomear_colunas)

# Normaliza SKU e preço da planilha
df_planilha["SKU"] = df_planilha["SKU"].apply(normalizar_sku)
df_planilha["Preco_Planilha"] = df_planilha["Preco_Planilha"].apply(normalizar_preco)

# Remove linhas sem SKU
df_planilha = df_planilha[df_planilha["SKU"].notna()].copy()

# Garante tipo numérico
df_planilha["Preco_Planilha"] = pd.to_numeric(
    df_planilha["Preco_Planilha"],
    errors="coerce"
)

# Identifica duplicados na planilha
df_duplicados = df_planilha[
    df_planilha.duplicated(subset=["SKU"], keep=False)
].copy()

if not df_duplicados.empty:
    df_duplicados["Status"] = "SKU duplicado na planilha"

# Mantém apenas o primeiro registro de cada SKU para comparação
df_planilha_base = df_planilha.drop_duplicates(
    subset=["SKU"],
    keep="first"
).copy()


# ============================================================
# CONSULTAR BANCO DE DADOS
# ============================================================

print("Conectando ao SQL Server...")

conexao = pyodbc.connect(
    f"DRIVER={{ODBC Driver 17 for SQL Server}};"
    f"SERVER={SERVIDOR};"
    f"DATABASE={BANCO};"
    f"UID={USUARIO};"
    f"PWD={SENHA};"
    f"TrustServerCertificate=yes;"
)

print("Executando consulta no banco...")

df_banco = pd.read_sql(QUERY_BANCO, conexao)

conexao.close()

validar_colunas(
    df_banco,
    ["SKU", "Preco_Banco"],
    "resultado da consulta SQL"
)

# Normaliza SKU e preço do banco
df_banco["SKU"] = df_banco["SKU"].apply(normalizar_sku)
df_banco["Preco_Banco"] = df_banco["Preco_Banco"].apply(normalizar_preco)

# Remove linhas sem SKU
df_banco = df_banco[df_banco["SKU"].notna()].copy()

# Garante tipo numérico
df_banco["Preco_Banco"] = pd.to_numeric(
    df_banco["Preco_Banco"],
    errors="coerce"
)

# Evita duplicidade do banco na comparação
df_banco_duplicados = df_banco[
    df_banco.duplicated(subset=["SKU"], keep=False)
].copy()

if not df_banco_duplicados.empty:
    df_banco_duplicados["Status"] = "SKU duplicado no banco"

df_banco_base = df_banco.drop_duplicates(
    subset=["SKU"],
    keep="first"
).copy()


# ============================================================
# COMPARAR PLANILHA X BANCO
# ============================================================

print("Comparando informações...")

df_comparacao = pd.merge(
    df_planilha_base,
    df_banco_base,
    on="SKU",
    how="outer",
    indicator=True
)

# Garante novamente que as colunas de preço sejam numéricas
df_comparacao["Preco_Planilha"] = pd.to_numeric(
    df_comparacao["Preco_Planilha"],
    errors="coerce"
)

df_comparacao["Preco_Banco"] = pd.to_numeric(
    df_comparacao["Preco_Banco"],
    errors="coerce"
)

# Classificação da linha
df_comparacao["Status"] = df_comparacao.apply(
    classificar_linha,
    axis=1
)

# Diferença em valor
df_comparacao["Diferenca_Valor"] = (
    df_comparacao["Preco_Planilha"] - df_comparacao["Preco_Banco"]
).round(2)

# Diferença percentual
df_comparacao["Diferenca_Percentual"] = np.nan

mascara_percentual = (
    df_comparacao["Preco_Banco"].notna()
    & (df_comparacao["Preco_Banco"] != 0)
    & df_comparacao["Preco_Planilha"].notna()
)

df_comparacao.loc[mascara_percentual, "Diferenca_Percentual"] = (
    (
        df_comparacao.loc[mascara_percentual, "Diferenca_Valor"]
        / df_comparacao.loc[mascara_percentual, "Preco_Banco"]
    ) * 100
).round(2)


# ============================================================
# ORGANIZAR RESULTADOS
# ============================================================

colunas_ordenadas = []

for coluna in [
    "SKU",
    "Produto_Planilha",
    "Produto_Banco",
    "Preco_Planilha",
    "Preco_Banco",
    "Diferenca_Valor",
    "Diferenca_Percentual",
    "Status"
]:
    if coluna in df_comparacao.columns:
        colunas_ordenadas.append(coluna)

outras_colunas = [
    coluna for coluna in df_comparacao.columns
    if coluna not in colunas_ordenadas and coluna != "_merge"
]

df_comparacao = df_comparacao[colunas_ordenadas + outras_colunas]

df_divergencias = df_comparacao[
    df_comparacao["Status"] != "Preço igual"
].copy()

df_precos_iguais = df_comparacao[
    df_comparacao["Status"] == "Preço igual"
].copy()

df_precos_divergentes = df_comparacao[
    df_comparacao["Status"] == "Preço divergente"
].copy()

df_existe_planilha_nao_banco = df_comparacao[
    df_comparacao["Status"] == "Produto existe na planilha, mas não existe no banco"
].copy()

df_existe_banco_nao_planilha = df_comparacao[
    df_comparacao["Status"] == "Produto existe no banco, mas não existe na planilha"
].copy()


# ============================================================
# RESUMO
# ============================================================

resumo = pd.DataFrame({
    "Indicador": [
        "Total de SKUs na planilha",
        "Total de SKUs no banco",
        "Total comparado",
        "Preços iguais",
        "Preços divergentes",
        "Existe na planilha, mas não existe no banco",
        "Existe no banco, mas não existe na planilha",
        "SKUs duplicados na planilha",
        "SKUs duplicados no banco"
    ],
    "Quantidade": [
        len(df_planilha_base),
        len(df_banco_base),
        len(df_comparacao),
        len(df_precos_iguais),
        len(df_precos_divergentes),
        len(df_existe_planilha_nao_banco),
        len(df_existe_banco_nao_planilha),
        len(df_duplicados),
        len(df_banco_duplicados)
    ]
})


# ============================================================
# GERAR EXCEL
# ============================================================

print("Gerando relatório Excel...")

with pd.ExcelWriter(CAMINHO_SAIDA, engine="openpyxl") as writer:
    resumo.to_excel(
        writer,
        sheet_name="Resumo",
        index=False
    )

    df_divergencias.to_excel(
        writer,
        sheet_name="Divergencias",
        index=False
    )

    df_precos_divergentes.to_excel(
        writer,
        sheet_name="Precos_Divergentes",
        index=False
    )

    df_precos_iguais.to_excel(
        writer,
        sheet_name="Precos_Iguais",
        index=False
    )

    df_existe_planilha_nao_banco.to_excel(
        writer,
        sheet_name="So_Na_Planilha",
        index=False
    )

    df_existe_banco_nao_planilha.to_excel(
        writer,
        sheet_name="So_No_Banco",
        index=False
    )

    df_duplicados.to_excel(
        writer,
        sheet_name="Duplicados_Planilha",
        index=False
    )

    df_banco_duplicados.to_excel(
        writer,
        sheet_name="Duplicados_Banco",
        index=False
    )

formatar_excel(CAMINHO_SAIDA)


# ============================================================
# FINALIZAÇÃO
# ============================================================

print("Comparação concluída com sucesso.")
print(f"Arquivo gerado em: {CAMINHO_SAIDA}")

print("\nResumo:")
print(resumo.to_string(index=False))